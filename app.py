# ============================================================
# PLAN TEJIDO v3 — Streamlit App
# ============================================================
# Ejecutar con: streamlit run app.py
# Instalar:     pip install streamlit pandas openpyxl plotly
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import re, math, io
from datetime import datetime, date
from fnmatch import fnmatchcase
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

st.set_page_config(
    page_title="Plan Tejido v3",
    page_icon="🧵",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# HELPERS
# ============================================================
def s(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return str(x).strip()

def collapse_spaces(t):
    return re.sub(r"\s+", " ", t).strip()

def norm_text(x):
    t = s(x)
    return collapse_spaces(t).upper() if t else ""

def norm_intlike(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        if float(x).is_integer():
            return str(int(float(x)))
        return collapse_spaces(str(x)).upper()
    t = collapse_spaces(str(x))
    if re.fullmatch(r"\d+\.0", t):
        return str(int(float(t)))
    return t.upper()

def machine(x, width=4):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    if isinstance(x, (int, np.integer)):
        t = str(int(x))
    elif isinstance(x, (float, np.floating)) and float(x).is_integer():
        t = str(int(float(x)))
    else:
        t = s(x)
        if re.fullmatch(r"\d+\.0", t):
            t = str(int(float(t)))
    t = t.strip()
    if t.isdigit() and len(t) < width:
        return t.zfill(width)
    return t

def lot_norm(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        if float(x).is_integer():
            return str(int(float(x)))
        return collapse_spaces(str(x)).upper()
    t = collapse_spaces(str(x))
    if re.fullmatch(r"\d+\.0", t):
        return str(int(float(t)))
    if "/" in t:
        a, b = t.split("/", 1)
        return a.strip().upper() + "/" + b.strip().upper()
    return t.upper()

def parse_date(x):
    if x is None or (isinstance(x, float) and math.isnan(x)) or s(x) == "":
        return pd.NaT
    if isinstance(x, (datetime, pd.Timestamp, date)):
        return pd.to_datetime(x).normalize()
    return pd.to_datetime(s(x), errors="coerce").normalize()

def split_allowed(val):
    t = norm_text(val)
    if not t:
        return []
    parts = re.split(r"[;,/|]+|\s{2,}", t)
    return [p.strip().upper() for p in parts if p.strip()]

def read_table(xls, sheet):
    return pd.read_excel(xls, sheet_name=sheet, header=1)

def read_optional(xls, sheet, cols):
    try:
        return read_table(xls, sheet)
    except Exception:
        return pd.DataFrame(columns=cols)

# ============================================================
# MOTOR PRINCIPAL — recibe todas las tablas ya cargadas
# ============================================================
def run_motor(
    params_tbl, reglas_tbl, estado_raw, demanda_raw,
    compat_raw, rates_raw, restr_raw, cal_raw,
    progress_cb=None
):
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    # ── PARAMETROS ──
    params = {}
    for _, r in params_tbl.iterrows():
        k = s(r.get("Campo"))
        if not k: continue
        v = r.get("Valor")
        if v is None or (isinstance(v, float) and math.isnan(v)) or s(v) == "":
            v = r.get("Ejemplo")
        params[k] = v

    def _float_param(name, default):
        t = s(params.get(name) or "")
        try:
            return float(t) if t else default
        except Exception:
            nums = re.findall(r"\d+(\.\d+)?", t)
            return float(nums[0]) if nums else default

    def _int_param(name, default):
        nums = re.findall(r"\d+", s(params.get(name) or ""))
        return int(nums[0]) if nums else default

    start_date   = parse_date(params.get("Fecha_inicio_plan"))
    end_date     = parse_date(params.get("Fecha_fin_plan"))
    hours_day    = _float_param("Horas_disponibles_dia", 24)
    min_lbs      = _float_param("Produccion_min_lbs", 100)
    pen_lote     = _float_param("Penalizacion_cambio_lote_horas", 4)
    pen_estilo   = _float_param("Penalizacion_cambio_estilo_horas", 8)
    pen_tejido   = _float_param("Penalizacion_cambio_tejido_horas", 24)
    rate_default = _float_param("Rate_default_lbs_dia", 1000)

    MAX_MAQ_POR_KEY         = _int_param("Max_maquinas_por_key", 9999)
    MIN_LBS_NUEVA_MAQ_KEY   = _float_param("Min_lbs_nueva_maquina_por_key", 2000)
    MIN_LBS_NUEVA_MAQ_RATIO = _float_param("Min_lbs_nueva_maquina_ratio", 0.25)
    MIN_HORAS_PROD_BF       = _float_param("Min_horas_produccion_backfill", 2)
    TAIL_MAX_LBS            = _float_param("Tail_max_lbs_key", 1000)
    TAIL_MIN_LBS_SEED       = _float_param("Tail_min_lbs_seed", 100)
    MAX_SEG_DIA             = _int_param("Max_segmentos_dia_normal", 2)
    MAX_SEG_DIA_TAIL        = _int_param("Max_segmentos_dia_tail", 2)

    if pd.isna(start_date) or pd.isna(end_date):
        raise ValueError("PARAMETROS: Fecha_inicio_plan y Fecha_fin_plan son obligatorias.")

    dates = pd.date_range(start_date, end_date, freq="D")
    N = len(dates)
    log(f"📅 Horizonte: {start_date.date()} → {end_date.date()} ({N} días)")

    def min_lbs_nueva_maquina(lbs_totales_key):
        return max(MIN_LBS_NUEVA_MAQ_KEY, lbs_totales_key * MIN_LBS_NUEVA_MAQ_RATIO)

    # REGLAS
    def rule_on(name, default=True):
        m = reglas_tbl.loc[
            reglas_tbl["Regla"].astype(str).str.strip().str.upper() == name.upper()
        ]
        if m.empty: return default
        v = norm_text(m.iloc[0].get("Valor (SI/NO)"))
        if v in ("SI","S","TRUE","1","YES","Y"): return True
        if v in ("NO","N","FALSE","0"):           return False
        return default

    permitir_setup = rule_on("Permitir_setup_si_hay_demanda", True)
    evitar_peq     = rule_on("Evitar_producciones_pequenas", True)
    usar_limite    = rule_on("Limite_maquinas_por_dia", False)

    max_machines = None
    _mmp = s(params.get("Max_maquinas_dia") or "")
    if usar_limite and _mmp:
        _nums = re.findall(r"\d+", _mmp)
        if _nums: max_machines = max(int(n) for n in _nums)

    # ── DEMANDA ──
    dem = demanda_raw.copy()
    _dem_required = ["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO","LBS_PENDIENTES"]
    for _c in _dem_required:
        if _c not in dem.columns:
            raise KeyError(f"DEMANDA: columna '{_c}' no encontrada.")

    dem["ESTILO_OPTIMO"]    = dem["ESTILO_OPTIMO"].apply(norm_text)
    dem["LOTE_HILO"]        = dem["LOTE_HILO"].apply(lot_norm)
    dem["DTITULAR"]         = dem["DTITULAR"].apply(norm_intlike)
    dem["TIPO_TEJIDO"]      = dem["TIPO_TEJIDO"].apply(norm_text)
    dem["YARN"]             = dem["YARN"].apply(norm_text)  if "YARN"  in dem.columns else ""
    dem["COLOR"]            = dem["COLOR"].apply(norm_text) if "COLOR" in dem.columns else ""
    dem["FECHA_COMPROMISO"] = dem["FECHA_COMPROMISO"].apply(parse_date) if "FECHA_COMPROMISO" in dem.columns else pd.NaT
    dem["LBS_PENDIENTES"]   = pd.to_numeric(dem["LBS_PENDIENTES"], errors="coerce").fillna(0.0).astype(float)
    prio_map = {"ALTA":3,"ALTO":3,"MEDIA":2,"MEDIO":2,"BAJA":1,"BAJO":1}
    dem["PRIO_NUM"] = dem["PRIORIDAD"].apply(lambda x: prio_map.get(norm_text(x), 2)) if "PRIORIDAD" in dem.columns else 2

    def due_bucket(due):
        if pd.isna(due): return 2
        if due < start_date: return 0
        if due <= end_date:  return 1
        return 2

    dem["_DUEB"] = dem["FECHA_COMPROMISO"].apply(due_bucket)
    dem["_DUED"] = dem["FECHA_COMPROMISO"].fillna(pd.Timestamp.max)

    dem_plan = (dem.groupby(["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO"], dropna=False)["LBS_PENDIENTES"]
                .sum().reset_index().rename(columns={"LBS_PENDIENTES":"LBS_PLAN"}))
    key_total_lbs = {(r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"]): float(r["LBS_PLAN"])
                     for _, r in dem_plan.iterrows()}
    key_due_map = {}
    for _, r in (dem.groupby(["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO"], dropna=False)["FECHA_COMPROMISO"]
                 .min().reset_index()).iterrows():
        key_due_map[(r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"])] = r["FECHA_COMPROMISO"]
    small_keys = {(r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"])
                  for _, r in dem_plan.iterrows() if r["LBS_PLAN"] <= TAIL_MAX_LBS}
    log(f"📦 Demanda: {len(dem)} filas, {dem['LBS_PENDIENTES'].sum():,.0f} LBS totales")

    # ── COMPAT_MAQUINA ──
    comp = compat_raw.copy()
    for _c in ["MAQUINA","TIPO_TEJIDO","DTITULAR"]:
        if _c not in comp.columns:
            raise KeyError(f"COMPAT_MAQUINA: columna '{_c}' no encontrada.")
    comp["MAQUINA"]     = comp["MAQUINA"].apply(machine)
    comp["TIPO_TEJIDO"] = comp["TIPO_TEJIDO"].apply(norm_text)
    comp["DTITULAR"]    = comp["DTITULAR"].apply(norm_intlike)
    comp["ACTIVA"]      = comp["ACTIVA"].apply(norm_text) if "ACTIVA" in comp.columns else "SI"
    compat_info = {}
    for _, r in comp.iterrows():
        maq = r["MAQUINA"]
        if not maq: continue
        if norm_text(r.get("ACTIVA")) in ("NO","N","0","FALSE"): continue
        tej_list = set(split_allowed(r.get("TIPO_TEJIDO")))
        tit      = norm_intlike(r.get("DTITULAR",""))
        if maq not in compat_info:
            compat_info[maq] = {"allowed": tej_list, "titular": tit}
        else:
            compat_info[maq]["allowed"].update(tej_list)
    if not compat_info:
        raise ValueError("COMPAT_MAQUINA: no hay máquinas activas.")
    log(f"🔧 Máquinas activas: {len(compat_info)}")

    # ── RESTRICCIONES ──
    restr = restr_raw.copy()
    _restr_cols = ["MAQUINA","ESTILO_OPTIMO","ESTILO_REAL","DTITULAR","DGREAL","TIPO_TEJIDO","PERMITIR"]
    for _c in _restr_cols:
        if _c not in restr.columns:
            raise KeyError(f"RESTRICCIONES: columna '{_c}' no encontrada.")
    restr["MAQUINA"]       = restr["MAQUINA"].apply(machine)
    restr["ESTILO_OPTIMO"] = restr["ESTILO_OPTIMO"].apply(norm_text)
    restr["ESTILO_REAL"]   = restr["ESTILO_REAL"].apply(norm_text)
    restr["DTITULAR"]      = restr["DTITULAR"].apply(norm_intlike)
    restr["DGREAL"]        = restr["DGREAL"].apply(norm_intlike)
    restr["TIPO_TEJIDO"]   = restr["TIPO_TEJIDO"].apply(norm_text)
    restr["PERMITIR"]      = restr["PERMITIR"].apply(norm_text)

    dgreal_by_tit = {}
    for _, r in restr.iterrows():
        tit = r["DTITULAR"]; dg = r["DGREAL"]
        if tit and dg:
            dgreal_by_tit.setdefault(tit, {})
            dgreal_by_tit[tit][dg] = dgreal_by_tit[tit].get(dg, 0) + 1
    dgreal_preferred = {tit: max(c, key=c.get) for tit, c in dgreal_by_tit.items()}

    estilo_optimo_to_real = {}
    for _, r in restr.iterrows():
        eo = r["ESTILO_OPTIMO"]; er = r["ESTILO_REAL"]
        if eo and er: estilo_optimo_to_real.setdefault(eo, er)

    def get_dgreal(maq, estilo_optimo, dtitular):
        m = machine(maq); eo = norm_text(estilo_optimo); dt = norm_intlike(dtitular)
        sub = restr[(restr["MAQUINA"]==m) & (restr["ESTILO_OPTIMO"]==eo) & (restr["DTITULAR"]==dt)]
        if not sub.empty:
            c = sub["DGREAL"].value_counts()
            if not c.empty: return c.index[0]
        sub2 = restr[(restr["ESTILO_OPTIMO"]==eo) & (restr["DTITULAR"]==dt)]
        if not sub2.empty:
            c2 = sub2["DGREAL"].value_counts()
            if not c2.empty: return c2.index[0]
        sub3 = restr[(restr["MAQUINA"]==m) & (restr["DTITULAR"]==dt)]
        if not sub3.empty:
            c3 = sub3["DGREAL"].value_counts()
            if not c3.empty: return c3.index[0]
        maq_tit = norm_intlike(compat_info.get(m, {}).get("titular", ""))
        if maq_tit: return maq_tit
        return dt

    def restriction_ok(maq, estilo_optimo, dtitular, tipo_tejido):
        m = machine(maq); eo = norm_text(estilo_optimo); dt = norm_intlike(dtitular)
        sub = restr[restr["MAQUINA"] == m]
        if sub.empty: return True
        sub = sub[(sub["ESTILO_OPTIMO"] == eo) | (sub["ESTILO_OPTIMO"] == "")]
        if sub.empty: return True
        sub_tit = sub[(sub["DTITULAR"] == dt) | (sub["DTITULAR"] == "")]
        if sub_tit.empty: return True
        for _, r in sub_tit.iterrows():
            if r["PERMITIR"].startswith("SI"): return True
        return False

    log(f"📋 Restricciones: {len(restr)} filas")

    # ── ESTADO_MAQUINA ──
    estado = estado_raw.copy()
    _estado_req = ["MAQUINA","FECHA_REF","ESTILO_OPTIMO","LOTE_HILO","DTITULAR",
                   "TIPO_TEJIDO","RATE_APROBADO","HORAS_LIBRES_REF"]
    for _c in _estado_req:
        if _c not in estado.columns:
            raise KeyError(f"ESTADO_MAQUINA: columna '{_c}' no encontrada.")
    estado["MAQUINA"]          = estado["MAQUINA"].apply(machine)
    estado["FECHA_REF"]        = estado["FECHA_REF"].apply(parse_date)
    estado["ESTILO_OPTIMO"]    = estado["ESTILO_OPTIMO"].apply(norm_text)
    estado["LOTE_HILO"]        = estado["LOTE_HILO"].apply(lot_norm)
    estado["DTITULAR"]         = estado["DTITULAR"].apply(norm_intlike)
    estado["TIPO_TEJIDO"]      = estado["TIPO_TEJIDO"].apply(norm_text)
    estado["COLOR"]            = estado["COLOR"].apply(norm_text) if "COLOR" in estado.columns else ""
    estado["YARN"]             = estado["YARN"].apply(norm_text) if "YARN"  in estado.columns else ""
    estado["RATE_APROBADO"]    = pd.to_numeric(estado["RATE_APROBADO"],    errors="coerce")
    estado["HORAS_LIBRES_REF"] = pd.to_numeric(estado["HORAS_LIBRES_REF"], errors="coerce")

    maq_state = {}
    for maq in compat_info.keys():
        maq_state[maq] = {"ESTILO_OPTIMO":"DISPONIBLE","LOTE_HILO":"DISPONIBLE",
                          "DTITULAR":compat_info[maq].get("titular",""),
                          "TIPO_TEJIDO":"","COLOR":"","YARN":""}
    for _, r in estado.iterrows():
        maq = r["MAQUINA"]
        if not maq or maq not in maq_state: continue
        for c in ["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO","COLOR","YARN"]:
            if c in estado.columns and s(r.get(c)) != "":
                maq_state[maq][c] = r.get(c)

    preplan_last_date  = {}; preplan_free_hours = {}
    preplan_active_day = [set() for _ in range(N)]
    for _, r in estado.iterrows():
        maq = r["MAQUINA"]
        if not maq or maq not in compat_info: continue
        fd = r.get("FECHA_REF"); hf = r.get("HORAS_LIBRES_REF")
        if pd.isna(fd) or pd.isna(hf): continue
        fd = pd.Timestamp(fd).normalize()
        try:    hf = float(hf)
        except: continue
        hf = max(0.0, min(hf, hours_day))
        preplan_last_date[maq]  = fd; preplan_free_hours[maq] = hf
        for di, dt in enumerate(dates):
            if pd.Timestamp(dt).normalize() <= fd:
                preplan_active_day[di].add(maq)

    # ── CALENDARIO ──
    cal = cal_raw.copy()
    hours_override = {}; blocked_day = set()
    if not cal.empty and "MAQUINA" in cal.columns:
        cal["MAQUINA"]      = cal["MAQUINA"].apply(machine)
        cal["FECHA_INICIO"] = cal["FECHA_INICIO"].apply(parse_date) if "FECHA_INICIO" in cal.columns else pd.NaT
        cal["FECHA_FIN"]    = cal["FECHA_FIN"].apply(parse_date)    if "FECHA_FIN"    in cal.columns else pd.NaT
        cal["TIPO"]         = cal["TIPO"].apply(norm_text) if "TIPO" in cal.columns else ""
        if "HORAS_DISPONIBLES" in cal.columns:
            cal["HORAS_DISPONIBLES"] = pd.to_numeric(cal["HORAS_DISPONIBLES"], errors="coerce")
        for _, r in cal.iterrows():
            maq = r.get("MAQUINA","")
            if not maq: continue
            fi = r.get("FECHA_INICIO"); ff = r.get("FECHA_FIN")
            if pd.isna(fi) or pd.isna(ff): continue
            tipo = norm_text(r.get("TIPO",""))
            horas = r.get("HORAS_DISPONIBLES")
            if isinstance(horas, float) and math.isnan(horas): horas = None
            if horas is None:
                if tipo in ("MANTENIMIENTO","PARO","STOP","DOWN"): horas = 0.0
                else: continue
            for di, dt in enumerate(dates):
                if fi <= dt <= ff:
                    if float(horas) <= 0:
                        blocked_day.add((maq, di)); hours_override[(maq, di)] = 0.0
                    else:
                        hours_override[(maq, di)] = float(horas)
    if preplan_last_date:
        date_to_index = {pd.Timestamp(d).normalize(): i for i, d in enumerate(dates)}
        for maq, last_dt in preplan_last_date.items():
            if last_dt not in date_to_index: continue
            last_i = date_to_index[last_dt]; free_h = float(preplan_free_hours.get(maq, 0.0))
            for di in range(0, last_i):
                blocked_day.add((maq, di)); hours_override[(maq, di)] = 0.0
            cal_h = float(hours_override.get((maq, last_i), hours_day))
            final_h = min(cal_h, free_h)
            if final_h <= 0:
                blocked_day.add((maq, last_i)); hours_override[(maq, last_i)] = 0.0
            else:
                hours_override[(maq, last_i)] = final_h

    def hours_avail(maq, di):
        return float(hours_override.get((maq, di), hours_day))

    # ── RATES ──
    rt = rates_raw.copy()
    for _c in ["MAQUINA","ESTILO_REAL","DTITULAR","TIPO_TEJIDO","LOTE_HILO","RATE_LBS_DIA"]:
        if _c not in rt.columns:
            raise KeyError(f"RATES: columna '{_c}' no encontrada.")
    rt["MAQUINA"]      = rt["MAQUINA"].apply(machine)
    rt["ESTILO_REAL"]  = rt["ESTILO_REAL"].apply(norm_text)
    rt["DTITULAR"]     = rt["DTITULAR"].apply(norm_intlike)
    rt["TIPO_TEJIDO"]  = rt["TIPO_TEJIDO"].apply(norm_text)
    rt["LOTE_HILO"]    = rt["LOTE_HILO"].apply(lot_norm)
    rt["TIPO_RATE"]    = rt["TIPO_RATE"].apply(norm_text) if "TIPO_RATE" in rt.columns else "RATE"
    rt["FUENTE"]       = rt["FUENTE"].apply(norm_text)    if "FUENTE"    in rt.columns else ""
    rt["RATE_LBS_DIA"] = pd.to_numeric(rt["RATE_LBS_DIA"], errors="coerce")
    rt = rt.dropna(subset=["RATE_LBS_DIA"]).copy()
    if "YARN" in rt.columns: rt["YARN"] = rt["YARN"].apply(norm_text)

    def tipo_rank(t):
        t = norm_text(t)
        if "APROB" in t:   return 0
        if "RATE" in t:    return 1
        if "HIST" in t:    return 2
        if "DEFAULT" in t: return 3
        return 9
    rt["_rank"] = rt["TIPO_RATE"].apply(tipo_rank)
    rt = rt.sort_values("_rank", ascending=True)

    _r_exact={}; _r_maq_est_dit_tej={}; _r_maq_est_tej={}
    _r_maq_tej={}; _r_default_tej={}; _r_global=None
    for _, r in rt.iterrows():
        maq=r["MAQUINA"]; est=r["ESTILO_REAL"]; dtit=r["DTITULAR"]
        tej=r["TIPO_TEJIDO"]; lot=r["LOTE_HILO"]; rate=float(r["RATE_LBS_DIA"])
        fuente = r.get("FUENTE") or r.get("TIPO_RATE") or "RATE"
        if maq and est and dtit and tej and lot:     _r_exact.setdefault((maq,est,dtit,tej,lot),(rate,fuente))
        if maq and est and dtit and tej and not lot: _r_maq_est_dit_tej.setdefault((maq,est,dtit,tej),(rate,fuente))
        if maq and est and tej and not dtit:         _r_maq_est_tej.setdefault((maq,est,tej),(rate,fuente))
        if maq and tej and not est and not dtit:     _r_maq_tej.setdefault((maq,tej),(rate,fuente))
        if not maq and tej and not est and not dtit: _r_default_tej.setdefault(tej,(rate,fuente))
        if not maq and not tej and not est and not dtit and _r_global is None: _r_global=(rate,fuente)

    def rate_lookup(maq, estilo_optimo, dtitular, tipo_tejido, lote_hilo):
        m=machine(maq); eo=norm_text(estilo_optimo); er=estilo_optimo_to_real.get(eo,eo)
        dt=norm_intlike(dtitular); tj=norm_text(tipo_tejido); lo=lot_norm(lote_hilo)
        _row = estado[(estado["MAQUINA"]==m) & (estado["ESTILO_OPTIMO"]==eo)]
        if not _row.empty:
            _ra = _row.iloc[0].get("RATE_APROBADO")
            if _ra and not pd.isna(_ra) and float(_ra)>0: return float(_ra),"RATE_APROBADO","ESTADO_MAQUINA"
        if (m,er,dt,tj,lo) in _r_exact:        r,f=_r_exact[(m,er,dt,tj,lo)];        return r,f,"MAQ+EST+DIT+TEJ+LOTE"
        if (m,er,dt,tj) in _r_maq_est_dit_tej: r,f=_r_maq_est_dit_tej[(m,er,dt,tj)]; return r,f,"MAQ+EST+DIT+TEJ"
        if (m,er,tj) in _r_maq_est_tej:        r,f=_r_maq_est_tej[(m,er,tj)];        return r,f,"MAQ+EST+TEJ"
        if (m,tj) in _r_maq_tej:               r,f=_r_maq_tej[(m,tj)];               return r,f,"MAQ+TEJ"
        if tj in _r_default_tej:               r,f=_r_default_tej[tj];               return r,f,"DEFAULT_TEJ"
        if _r_global is not None:              r,f=_r_global;                         return r,f,"DEFAULT_GLOBAL"
        return rate_default,"DEFAULT_PARAM","DEFAULT_PARAM"

    # ── MOTOR ──
    schedule            = {maq: [[] for _ in range(N)] for maq in compat_info.keys()}
    machines_active_day = [set(preplan_active_day[di]) for di in range(N)]
    seg_counter         = {maq: 0 for maq in compat_info.keys()}

    def dkey(eo, lote, dtit, tej):
        return (norm_text(eo), lot_norm(lote), norm_intlike(dtit), norm_text(tej))

    key_to_idxs = {}
    for idx, row in dem.iterrows():
        k = dkey(row["ESTILO_OPTIMO"],row["LOTE_HILO"],row["DTITULAR"],row["TIPO_TEJIDO"])
        key_to_idxs.setdefault(k, []).append(idx)

    def can_use_day(maq, di):
        if (maq,di) in blocked_day:        return False
        if hours_avail(maq,di) <= 0:       return False
        if max_machines is None:            return True
        if maq in machines_active_day[di]: return True
        return len(machines_active_day[di]) < max_machines

    def mark_active(maq, di):
        machines_active_day[di].add(maq)

    def machine_can(maq, estilo_optimo, dtitular, tipo_tejido, lote_hilo=None):
        info = compat_info[maq]
        tit_fix = norm_intlike(info.get("titular",""))
        dt      = norm_intlike(dtitular)
        if tit_fix and dt and tit_fix != dt: return False
        allowed = info.get("allowed", set())
        tj      = norm_text(tipo_tejido)
        if allowed and tj and tj not in allowed: return False
        if not restriction_ok(maq, estilo_optimo, dtitular, tipo_tejido): return False
        return True

    def prev_last_key_before_day(maq, i):
        for j in range(i-1, -1, -1):
            if schedule[maq][j]: return schedule[maq][j][-1]
        return maq_state[maq]

    def penalty(prevk, newk):
        prev_est = norm_text(prevk.get("ESTILO_OPTIMO",""))
        if prev_est in ("","DISPONIBLE"): return 0.0, "INICIO"
        ptj = norm_text(prevk.get("TIPO_TEJIDO","")); ntj = norm_text(newk.get("TIPO_TEJIDO",""))
        if ptj and ntj and ptj != ntj: return float(pen_tejido), "CAMBIO_TEJIDO"
        if prev_est != norm_text(newk.get("ESTILO_OPTIMO","")): return float(pen_estilo), "CAMBIO_ESTILO"
        if lot_norm(prevk.get("LOTE_HILO","")) != lot_norm(newk.get("LOTE_HILO","")): return float(pen_lote), "CAMBIO_LOTE"
        return 0.0, "CONTINUIDAD"

    def day_used_hours(maq, di):
        local_h = hours_avail(maq, di)
        if local_h <= 0: return 0.0
        used = 0.0
        for seg in schedule[maq][di]:
            used += float(seg["HORAS_SETUP"])
            rate = float(seg["RATE_LBS_DIA"]); lbs = float(seg["LBS_ASIGNADAS"])
            used += (lbs / rate) * local_h if rate > 0 else 0.0
        return used

    def remaining_hours(maq, di):
        return max(0.0, hours_avail(maq, di) - day_used_hours(maq, di))

    def capacity_from_hours(rate, local_h, prod_h):
        if local_h <= 0: return 0.0
        return float(rate) * (float(prod_h) / float(local_h))

    def last_assigned_state(maq):
        last_seg = None; last_seq = -1
        for di in range(N):
            for seg in schedule[maq][di]:
                if seg["SEQ_MAQUINA"] > last_seq:
                    last_seq = seg["SEQ_MAQUINA"]; last_seg = seg
        return last_seg

    def affinity_score(maq, item):
        last = last_assigned_state(maq)
        if last is None:
            st = maq_state.get(maq, {})
            cur_estilo = norm_text(st.get("ESTILO_OPTIMO",""))
            if cur_estilo in ("","DISPONIBLE"): return 1
            cur_lote  = lot_norm(st.get("LOTE_HILO",""))
            cur_yarn  = norm_text(st.get("YARN",""))
            cur_color = norm_text(st.get("COLOR",""))
        else:
            cur_estilo = norm_text(last.get("ESTILO_OPTIMO",""))
            cur_lote   = lot_norm(last.get("LOTE_HILO",""))
            cur_yarn   = norm_text(last.get("YARN",""))
            cur_color  = norm_text(last.get("COLOR",""))
        new_estilo = norm_text(item.get("ESTILO_OPTIMO","") if isinstance(item,dict) else item["ESTILO_OPTIMO"])
        new_lote   = lot_norm(item.get("LOTE_HILO","")     if isinstance(item,dict) else item["LOTE_HILO"])
        new_yarn   = norm_text(item.get("YARN","")         if isinstance(item,dict) else item["YARN"])
        new_color  = norm_text(item.get("COLOR","")        if isinstance(item,dict) else item["COLOR"])
        if cur_estilo == new_estilo and cur_lote == new_lote: return 4
        if cur_yarn  and new_yarn  and cur_yarn  == new_yarn:  return 3
        if cur_color and new_color and cur_color == new_color: return 2
        return 1

    def add_segment(maq, di, keydict, lbs, setup_h, setup_tipo):
        rate, fuente, match = rate_lookup(maq, keydict["ESTILO_OPTIMO"], keydict["DTITULAR"],
                                          keydict["TIPO_TEJIDO"], keydict["LOTE_HILO"])
        eo = keydict["ESTILO_OPTIMO"]; dt = keydict["DTITULAR"]
        er = estilo_optimo_to_real.get(norm_text(eo), norm_text(eo))
        dg = keydict.get("DGREAL") or get_dgreal(maq, eo, dt)
        seg_counter[maq] += 1
        seg = {"ESTILO_OPTIMO":eo,"ESTILO_REAL":er,"LOTE_HILO":keydict["LOTE_HILO"],
               "DTITULAR":dt,"DGREAL":dg,"TIPO_TEJIDO":keydict["TIPO_TEJIDO"],
               "YARN":keydict.get("YARN",""),"COLOR":keydict.get("COLOR",""),
               "LBS_ASIGNADAS":float(lbs),"HORAS_SETUP":float(setup_h),
               "RATE_LBS_DIA":float(rate),"FUENTE_RATE":fuente,"RATE_MATCH":match,
               "SETUP_TIPO":setup_tipo,"SEQ_MAQUINA":seg_counter[maq]}
        schedule[maq][di].append(seg)
        mark_active(maq, di)
        maq_state[maq]["ESTILO_OPTIMO"] = seg["ESTILO_OPTIMO"]
        maq_state[maq]["LOTE_HILO"]     = seg["LOTE_HILO"]
        maq_state[maq]["TIPO_TEJIDO"]   = seg["TIPO_TEJIDO"]
        maq_state[maq]["YARN"]          = seg.get("YARN","")
        maq_state[maq]["COLOR"]         = seg.get("COLOR","")

    def earliest_free_day(maq):
        for i in range(N):
            if can_use_day(maq, i) and len(schedule[maq][i]) == 0: return i
        return None

    def machines_used_for_key(keydict):
        ms = set()
        eo=keydict["ESTILO_OPTIMO"]; lo=keydict["LOTE_HILO"]
        dt=keydict["DTITULAR"];      tj=keydict["TIPO_TEJIDO"]
        for maq in schedule.keys():
            for di in range(N):
                for seg in schedule[maq][di]:
                    if (seg["ESTILO_OPTIMO"],seg["LOTE_HILO"],seg["DTITULAR"],seg["TIPO_TEJIDO"])==(eo,lo,dt,tj):
                        if float(seg["LBS_ASIGNADAS"]) > 1e-9:
                            ms.add(maq); break
        return ms

    def can_place_due(keyk, di):
        due = key_due_map.get(keyk, pd.NaT)
        if pd.isna(due): return True
        return dates[di] <= due

    def _make_keydict(drow, maq=""):
        eo = norm_text(drow.get("ESTILO_OPTIMO","") if isinstance(drow,dict) else drow["ESTILO_OPTIMO"])
        dt = norm_intlike(drow.get("DTITULAR","")   if isinstance(drow,dict) else drow["DTITULAR"])
        lo = lot_norm(drow.get("LOTE_HILO","")      if isinstance(drow,dict) else drow["LOTE_HILO"])
        tj = norm_text(drow.get("TIPO_TEJIDO","")   if isinstance(drow,dict) else drow["TIPO_TEJIDO"])
        yn = norm_text(drow.get("YARN",""))
        cl = norm_text(drow.get("COLOR",""))
        return {"ESTILO_OPTIMO":eo,"LOTE_HILO":lo,"DTITULAR":dt,"TIPO_TEJIDO":tj,
                "YARN":yn,"COLOR":cl,"DGREAL":get_dgreal(maq, eo, dt)}

    def assign_primary_block(maq, idx, start_i):
        remaining = float(dem.at[idx, "LBS_PENDIENTES"])
        if remaining <= 0 or start_i is None or start_i >= N: return remaining
        drow = dem.loc[idx]
        if not machine_can(maq, drow["ESTILO_OPTIMO"], drow["DTITULAR"], drow["TIPO_TEJIDO"], drow["LOTE_HILO"]):
            return remaining
        keydict = _make_keydict(drow, maq)
        i = start_i; first = True
        while i < N and float(dem.at[idx, "LBS_PENDIENTES"]) > 1e-9:
            if not can_use_day(maq, i): break
            if len(schedule[maq][i]) != 0: break
            local_h = hours_avail(maq, i)
            prev = prev_last_key_before_day(maq, i)
            sh, st = penalty(prev, keydict)
            if not permitir_setup and sh > 0: break
            horas_netas = max(0.0, local_h - sh)
            if horas_netas <= 0: break
            rate, _, _ = rate_lookup(maq, keydict["ESTILO_OPTIMO"], keydict["DTITULAR"],
                                     keydict["TIPO_TEJIDO"], keydict["LOTE_HILO"])
            cap = rate * (horas_netas / local_h)
            lbs = min(float(dem.at[idx, "LBS_PENDIENTES"]), cap)
            if evitar_peq and lbs > 0 and lbs < min_lbs and float(dem.at[idx,"LBS_PENDIENTES"]) >= min_lbs: break
            if lbs <= 0: break
            add_segment(maq, i, keydict, lbs, sh, ("INICIO" if first else "CONTINUIDAD") if st=="CONTINUIDAD" else st)
            dem.at[idx, "LBS_PENDIENTES"] = max(0.0, float(dem.at[idx,"LBS_PENDIENTES"]) - lbs)
            first = False; i += 1
        return float(dem.at[idx, "LBS_PENDIENTES"])

    # FASE 0
    log("⚙️ Fase 0: continuidad heredada...")
    for maq, st in maq_state.items():
        eo = norm_text(st.get("ESTILO_OPTIMO",""))
        if eo not in ("","DISPONIBLE"):
            k = dkey(eo, st.get("LOTE_HILO",""), st.get("DTITULAR",""), st.get("TIPO_TEJIDO",""))
            idxs = key_to_idxs.get(k, [])
            if idxs:
                si = earliest_free_day(maq)
                if si is not None:
                    lbs_rem   = float(dem.at[idxs[0], "LBS_PENDIENTES"])
                    total_key = key_total_lbs.get(k, lbs_rem)
                    umbral    = min_lbs_nueva_maquina(total_key)
                    maqs_ya   = machines_used_for_key({"ESTILO_OPTIMO":k[0],"LOTE_HILO":k[1],"DTITULAR":k[2],"TIPO_TEJIDO":k[3]})
                    maqs_compat = [m for m in compat_info if machine_can(m, k[0], k[2], k[3])]
                    es_unica = len(maqs_compat) <= 1
                    if len(maqs_ya) == 0 or lbs_rem >= umbral or es_unica:
                        assign_primary_block(maq, idxs[0], si)

    # FASE 1
    log("⚙️ Fase 1: asignación primaria...")
    rem = dem[dem["LBS_PENDIENTES"] > 0].copy()
    groups = list(rem.groupby(["DTITULAR","TIPO_TEJIDO"]).groups.keys())
    group_scores = []
    for g in groups:
        sub        = rem[(rem["DTITULAR"]==g[0]) & (rem["TIPO_TEJIDO"]==g[1])]
        dueb       = int(sub["_DUEB"].min()) if not sub.empty else 2
        tot        = float(sub["LBS_PENDIENTES"].sum())
        tit        = norm_intlike(g[0])
        dg_defined = 1 if tit in dgreal_preferred else 0
        group_scores.append((g, dueb, -tot, -dg_defined))
    group_scores.sort(key=lambda x: (x[1], x[2], x[3], x[0]))

    for (g, _, __, ___) in group_scores:
        tit_g, tej_g = g
        sub_idx = dem[(dem["DTITULAR"]==tit_g) & (dem["TIPO_TEJIDO"]==tej_g) & (dem["LBS_PENDIENTES"]>0)].index.tolist()
        sub_idx.sort(key=lambda i: (dem.at[i,"_DUEB"],dem.at[i,"_DUED"],-dem.at[i,"PRIO_NUM"],
                                    -float(dem.at[i,"LBS_PENDIENTES"]),dem.at[i,"ESTILO_OPTIMO"],dem.at[i,"LOTE_HILO"]))
        for idx in sub_idx:
            if float(dem.at[idx,"LBS_PENDIENTES"]) <= 0: continue
            drow = dem.loc[idx]
            keyk = (drow["ESTILO_OPTIMO"],drow["LOTE_HILO"],drow["DTITULAR"],drow["TIPO_TEJIDO"])
            if keyk in small_keys: continue
            total_key_lbs    = key_total_lbs.get(keyk, float(dem.at[idx,"LBS_PENDIENTES"]))
            umbral_nueva_maq = min_lbs_nueva_maquina(total_key_lbs)
            used = list(machines_used_for_key({"ESTILO_OPTIMO":drow["ESTILO_OPTIMO"],"LOTE_HILO":drow["LOTE_HILO"],
                                               "DTITULAR":drow["DTITULAR"],"TIPO_TEJIDO":drow["TIPO_TEJIDO"]}))
            while float(dem.at[idx,"LBS_PENDIENTES"]) > 1e-9:
                remaining = float(dem.at[idx,"LBS_PENDIENTES"]); candidates = []
                for maq in schedule.keys():
                    open_new = maq not in used
                    if open_new:
                        if len(used) >= MAX_MAQ_POR_KEY: continue
                        if remaining < umbral_nueva_maq: continue
                    si = earliest_free_day(maq)
                    if si is None: continue
                    if not machine_can(maq, drow["ESTILO_OPTIMO"], drow["DTITULAR"], drow["TIPO_TEJIDO"], drow["LOTE_HILO"]): continue
                    keydict_c = _make_keydict(drow, maq)
                    total_cap = 0.0; i = si; first_c = True
                    while i < N and len(schedule[maq][i]) == 0 and can_use_day(maq, i):
                        local_h = hours_avail(maq, i)
                        prev = prev_last_key_before_day(maq, i)
                        sh, _ = penalty(prev, keydict_c) if first_c else (0.0,"CONTINUIDAD")
                        if not permitir_setup and sh > 0: break
                        hn = max(0.0, local_h - sh)
                        if hn <= 0: break
                        rt_v,_,_ = rate_lookup(maq, keydict_c["ESTILO_OPTIMO"], keydict_c["DTITULAR"],
                                               keydict_c["TIPO_TEJIDO"], keydict_c["LOTE_HILO"])
                        total_cap += rt_v * (hn / local_h); first_c = False; i += 1
                    prev0 = prev_last_key_before_day(maq, si)
                    sh0, _ = penalty(prev0, keydict_c)
                    usedflag = 1 if maq in used else 0
                    aff = affinity_score(maq, keydict_c)
                    candidates.append((maq, usedflag, aff, total_cap, sh0, si))
                if not candidates: break
                candidates.sort(key=lambda x: (-x[1], -x[2], -x[3], x[4], x[0]))
                maq_best, _, _aff, __cap, __sh0, si_best = candidates[0]
                before = float(dem.at[idx,"LBS_PENDIENTES"])
                assign_primary_block(maq_best, idx, si_best)
                after = float(dem.at[idx,"LBS_PENDIENTES"])
                if after >= before - 1e-9: break
                if maq_best not in used: used.append(maq_best)

    # FASE 2
    log("⚙️ Fase 2: backfill...")
    def future_min_affinity(maq, di):
        min_aff = 1; found = False
        for dj in range(di + 1, N):
            for seg in schedule[maq][dj]:
                if float(seg.get("LBS_ASIGNADAS", 0)) > 1e-9:
                    aff_future = affinity_score(maq, seg)
                    if not found or aff_future < min_aff:
                        min_aff = aff_future
                    found = True
        return min_aff if found else 1

    def backfill_one_slot(maq, di, max_segments):
        if not can_use_day(maq, di):              return False
        if len(schedule[maq][di]) == 0:            return False
        if len(schedule[maq][di]) >= max_segments: return False
        free_h = remaining_hours(maq, di)
        if free_h <= 1e-9:                         return False
        prevk   = schedule[maq][di][-1]
        local_h = hours_avail(maq, di)
        cand_idxs = dem[dem["LBS_PENDIENTES"] > 1e-9].index.tolist()
        if not cand_idxs: return False
        min_future_aff = future_min_affinity(maq, di)
        aff_cache = {i: affinity_score(maq, dem.loc[i]) for i in cand_idxs}
        cand_idxs.sort(key=lambda i: (-aff_cache[i], dem.at[i,"_DUEB"], dem.at[i,"_DUED"],
                                      -dem.at[i,"PRIO_NUM"], -float(dem.at[i,"LBS_PENDIENTES"])))
        for idx in cand_idxs:
            r    = dem.loc[idx]
            keyk = (r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"])
            if keyk in small_keys: continue
            keydict = _make_keydict(r, maq)
            if not machine_can(maq, keydict["ESTILO_OPTIMO"], keydict["DTITULAR"],
                               keydict["TIPO_TEJIDO"], keydict["LOTE_HILO"]): continue
            sh2, st2 = penalty(prevk, keydict)
            aff = aff_cache[idx]
            if aff < min_future_aff: continue
            if (not permitir_setup) and sh2 > 0 and aff < 2: continue
            if free_h < sh2 + MIN_HORAS_PROD_BF: continue
            prod_h = max(0.0, free_h - sh2)
            rate2,_,_ = rate_lookup(maq, keydict["ESTILO_OPTIMO"], keydict["DTITULAR"],
                                    keydict["TIPO_TEJIDO"], keydict["LOTE_HILO"])
            cap2 = capacity_from_hours(rate2, local_h, prod_h)
            rem2 = float(dem.at[idx,"LBS_PENDIENTES"])
            lbs2 = min(rem2, cap2)
            if evitar_peq and lbs2 > 0 and lbs2 < min_lbs and rem2 >= min_lbs: continue
            if lbs2 <= 1e-9: continue
            add_segment(maq, di, keydict, lbs2, sh2, st2)
            dem.at[idx,"LBS_PENDIENTES"] = max(0.0, rem2 - lbs2)
            return True
        return False

    for di in range(N):
        for maq in schedule.keys():
            backfill_one_slot(maq, di, MAX_SEG_DIA)

    # FASE 2.5
    log("⚙️ Fase 2.5: tails...")
    def seed_and_continue():
        small_idxs = [idx for idx, r in dem.iterrows()
                      if (r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"]) in small_keys
                      and float(r["LBS_PENDIENTES"]) > 1e-9]
        small_idxs.sort(key=lambda i: (dem.at[i,"_DUEB"],dem.at[i,"_DUED"],
                                       -dem.at[i,"PRIO_NUM"],-float(dem.at[i,"LBS_PENDIENTES"])))
        for idx in small_idxs:
            if float(dem.at[idx,"LBS_PENDIENTES"]) <= 1e-9: continue
            r    = dem.loc[idx]
            keyk = (r["ESTILO_OPTIMO"],r["LOTE_HILO"],r["DTITULAR"],r["TIPO_TEJIDO"])
            remaining = float(dem.at[idx,"LBS_PENDIENTES"])
            for di in range(N-1):
                if remaining <= 1e-9: break
                if not can_place_due(keyk,di) or not can_place_due(keyk,di+1): continue
                for maq in schedule.keys():
                    if remaining <= 1e-9: break
                    if len(schedule[maq][di]) == 0: continue
                    if len(schedule[maq][di]) >= MAX_SEG_DIA_TAIL: continue
                    if len(schedule[maq][di+1]) != 0: continue
                    if not can_use_day(maq,di) or not can_use_day(maq,di+1): continue
                    keydict = _make_keydict(r, maq)
                    if not machine_can(maq,keydict["ESTILO_OPTIMO"],keydict["DTITULAR"],
                                       keydict["TIPO_TEJIDO"],keydict["LOTE_HILO"]): continue
                    free_h = remaining_hours(maq, di)
                    if free_h <= 1e-9: continue
                    prevk = schedule[maq][di][-1]
                    sh, st = penalty(prevk, keydict)
                    if not permitir_setup and sh > 0: continue
                    if free_h < sh + MIN_HORAS_PROD_BF: continue
                    local_h  = hours_avail(maq, di)
                    rate,_,_ = rate_lookup(maq,keydict["ESTILO_OPTIMO"],keydict["DTITULAR"],
                                           keydict["TIPO_TEJIDO"],keydict["LOTE_HILO"])
                    cap_di   = capacity_from_hours(rate, local_h, max(0.0, free_h-sh))
                    local_h2 = hours_avail(maq, di+1)
                    cap_di1  = capacity_from_hours(rate, local_h2, local_h2)
                    seed_lbs = min(remaining, cap_di)
                    if seed_lbs < TAIL_MIN_LBS_SEED and remaining > TAIL_MIN_LBS_SEED: continue
                    rem_after = remaining - seed_lbs
                    cont_lbs  = min(rem_after, cap_di1)
                    if cont_lbs <= 1e-9 and rem_after > 1e-6: continue
                    add_segment(maq,di,keydict,seed_lbs,sh,"TAIL_SEED" if st=="CONTINUIDAD" else st)
                    if cont_lbs > 1e-9: add_segment(maq,di+1,keydict,cont_lbs,0.0,"TAIL_CONT")
                    remaining -= (seed_lbs + cont_lbs)
                    dem.at[idx,"LBS_PENDIENTES"] = max(0.0, remaining)
    seed_and_continue()

    # FASE 3
    log("⚙️ Fase 3: máquinas ociosas...")
    def machine_has_any_lbs(maq):
        for di in range(N):
            for seg in schedule[maq][di]:
                if float(seg.get("LBS_ASIGNADAS",0.0)) > 1e-9: return True
        return False

    def best_pending_for_machine(maq):
        info = compat_info[maq]
        tit_fix = norm_intlike(info.get("titular",""))
        allowed = info.get("allowed", set())
        cand = dem[dem["LBS_PENDIENTES"] > 1e-9].copy()
        if cand.empty: return None
        if tit_fix != "": cand = cand[cand["DTITULAR"] == tit_fix]
        if allowed:       cand = cand[cand["TIPO_TEJIDO"].isin(allowed)]
        if cand.empty: return None
        cand = cand.sort_values(by=["_DUEB","_DUED","PRIO_NUM","LBS_PENDIENTES"], ascending=[True,True,False,False])
        for idx in cand.index.tolist():
            r = dem.loc[idx]
            if machine_can(maq, r["ESTILO_OPTIMO"], r["DTITULAR"], r["TIPO_TEJIDO"], r["LOTE_HILO"]): return idx
        return None

    passes = 0
    while passes < 6:
        passes += 1
        idle_machines = [m for m in schedule.keys() if not machine_has_any_lbs(m)]
        if not idle_machines: break
        if not (dem["LBS_PENDIENTES"] > 1e-9).any(): break
        changed = False
        for maq in idle_machines:
            si = earliest_free_day(maq); idx = best_pending_for_machine(maq)
            if si is None or idx is None: continue
            before = float(dem.at[idx,"LBS_PENDIENTES"])
            assign_primary_block(maq, idx, si)
            after = float(dem.at[idx,"LBS_PENDIENTES"])
            if after < before - 1e-9: changed = True
        if not changed: break

    lbs_pend_fin = float(dem["LBS_PENDIENTES"].sum())
    log(f"✅ Motor finalizado. LBS pendientes: {lbs_pend_fin:,.0f}")

    # ── BUILD OUTPUT ──
    log("📊 Construyendo tablas de salida...")
    plan_rows = []
    for maq in schedule.keys():
        for di in range(N):
            for seg in schedule[maq][di]:
                local_h     = hours_avail(maq, di)
                horas_netas = max(0.0, local_h - float(seg["HORAS_SETUP"]))
                plan_rows.append({
                    "FECHA":dates[di].date(),"DIA":di+1,"MAQUINA":maq,
                    "SEQ_MAQUINA":seg["SEQ_MAQUINA"],
                    "ESTILO_OPTIMO":seg["ESTILO_OPTIMO"],"ESTILO_REAL":seg.get("ESTILO_REAL",""),
                    "LOTE_HILO":seg["LOTE_HILO"],"DTITULAR":seg["DTITULAR"],
                    "DGREAL":seg.get("DGREAL",""),"TIPO_TEJIDO":seg["TIPO_TEJIDO"],
                    "YARN":seg.get("YARN",""),"COLOR":seg.get("COLOR",""),
                    "LBS_ASIGNADAS":float(seg["LBS_ASIGNADAS"]),"HORAS_SETUP":float(seg["HORAS_SETUP"]),
                    "HORAS_NETAS":float(horas_netas),"RATE_LBS_DIA":float(seg["RATE_LBS_DIA"]),
                    "FUENTE_RATE":seg["FUENTE_RATE"],"RATE_MATCH":seg["RATE_MATCH"],
                    "SETUP_TIPO":seg["SETUP_TIPO"],
                })

    plan = pd.DataFrame(plan_rows)
    free_rows = []
    for maq in schedule.keys():
        for di, dt in enumerate(dates):
            if len(schedule[maq][di]) == 0:
                tipo = "BLOQUEADA" if (maq,di) in blocked_day else "LIBRE"
                if tipo == "LIBRE" and max_machines is not None:
                    if maq not in machines_active_day[di] and len(machines_active_day[di]) >= max_machines:
                        tipo = "CAP_DIA"
                free_rows.append({
                    "FECHA":dates[di].date(),"DIA":di+1,"MAQUINA":maq,"SEQ_MAQUINA":0,
                    "ESTILO_OPTIMO":"","ESTILO_REAL":"","LOTE_HILO":"","DTITULAR":"",
                    "DGREAL":"","TIPO_TEJIDO":"","YARN":"","COLOR":"",
                    "LBS_ASIGNADAS":0.0,"HORAS_SETUP":0.0,"HORAS_NETAS":0.0,
                    "RATE_LBS_DIA":0.0,"FUENTE_RATE":"","RATE_MATCH":"","SETUP_TIPO":tipo,
                })

    plan_full = pd.concat([plan, pd.DataFrame(free_rows)], ignore_index=True)
    plan_full["MAQUINA"]   = plan_full["MAQUINA"].astype(str).str.replace(r"\.0$","",regex=True).str.zfill(4)
    plan_full["DTITULAR"]  = plan_full["DTITULAR"].astype(str).str.replace(r"\.0$","",regex=True)
    plan_full["LOTE_HILO"] = plan_full["LOTE_HILO"].astype(str)
    plan_full = plan_full.sort_values(["MAQUINA","SEQ_MAQUINA","FECHA","DIA"]).reset_index(drop=True)

    resumen_asig = (plan_full[plan_full["LBS_ASIGNADAS"]>0]
                    .groupby(["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO"],dropna=False)["LBS_ASIGNADAS"]
                    .sum().reset_index())
    resumen = dem_plan.merge(resumen_asig, on=["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO"], how="outer")
    resumen["LBS_PLAN"]      = pd.to_numeric(resumen["LBS_PLAN"],      errors="coerce").fillna(0.0)
    resumen["LBS_ASIGNADAS"] = pd.to_numeric(resumen["LBS_ASIGNADAS"], errors="coerce").fillna(0.0)
    resumen["DIFERENCIA"]    = resumen["LBS_PLAN"] - resumen["LBS_ASIGNADAS"]
    resumen["EXCEDIDO"]      = np.where(resumen["LBS_ASIGNADAS"] > resumen["LBS_PLAN"]+1e-6,"SI","NO")
    resumen = resumen.sort_values(["EXCEDIDO","LBS_ASIGNADAS"], ascending=[False,False])

    md_rows = []
    for di, dt in enumerate(dates):
        day_plan   = plan_full[(plan_full["DIA"]==di+1) & (plan_full["LBS_ASIGNADAS"]>0)]
        maqs_motor = set(day_plan["MAQUINA"].unique())
        maqs_pre   = {machine(x) for x in preplan_active_day[di]}
        md_rows.append({"FECHA":dt.date(),"DIA":di+1,
                        "MAQUINAS_ACTIVAS":len(maqs_motor|maqs_pre),
                        "MAQUINAS_MOTOR":len(maqs_motor),"MAQUINAS_PREPLAN":len(maqs_pre),
                        "LBS_TOTALES":float(day_plan["LBS_ASIGNADAS"].sum()),
                        "HORAS_SETUP_TOTALES":float(day_plan["HORAS_SETUP"].sum())})
    machines_day_df = pd.DataFrame(md_rows)

    pivot_src = plan_full[plan_full["LBS_ASIGNADAS"]>0].copy()
    pivot_src["FECHA"] = pd.to_datetime(pivot_src["FECHA"])
    pivot_plan = (pivot_src.pivot_table(
        index=["TIPO_TEJIDO","MAQUINA","ESTILO_OPTIMO","ESTILO_REAL",
               "LOTE_HILO","DTITULAR","DGREAL","YARN","COLOR"],
        columns="FECHA", values="LBS_ASIGNADAS", aggfunc="sum", fill_value=0.0)
        .reset_index())
    date_cols = sorted([c for c in pivot_plan.columns if isinstance(c, (pd.Timestamp, datetime))])

    def first_prod_date(row):
        for c in date_cols:
            if row[c] > 0: return c
        return pd.NaT

    pivot_plan["_FI"] = pivot_plan.apply(first_prod_date, axis=1)
    pivot_plan = pivot_plan.sort_values(["_FI","TIPO_TEJIDO","MAQUINA","ESTILO_OPTIMO","LOTE_HILO","DTITULAR"])
    pivot_plan = pivot_plan[["TIPO_TEJIDO","MAQUINA","ESTILO_OPTIMO","ESTILO_REAL",
                              "DTITULAR","DGREAL","YARN","COLOR","LOTE_HILO"] + date_cols]
    pivot_plan["Total_general"] = pivot_plan[date_cols].sum(axis=1)
    pivot_plan = pivot_plan.rename(columns={c: c.strftime("%Y-%m-%d") for c in date_cols}).reset_index(drop=True)

    total_assigned = float(plan_full["LBS_ASIGNADAS"].sum())
    total_setup    = float(plan_full["HORAS_SETUP"].sum())
    prod_hours     = float(plan_full.loc[plan_full["LBS_ASIGNADAS"]>0,"HORAS_NETAS"].sum())

    kpi = pd.DataFrame([
        {"KPI":"LBS_TOTALES_PLANIFICADAS",  "VALOR":total_assigned},
        {"KPI":"LBS_PENDIENTES_RESTANTES",  "VALOR":lbs_pend_fin},
        {"KPI":"PCT_COBERTURA",             "VALOR":round(total_assigned/(total_assigned+lbs_pend_fin)*100,2) if (total_assigned+lbs_pend_fin)>0 else 0},
        {"KPI":"HORAS_SETUP_TOTALES",       "VALOR":total_setup},
        {"KPI":"HORAS_PRODUCTIVAS_TOTALES", "VALOR":prod_hours},
        {"KPI":"MIN_LBS_NUEVA_MAQ_KEY_ABS", "VALOR":MIN_LBS_NUEVA_MAQ_KEY},
        {"KPI":"MIN_LBS_NUEVA_MAQ_RATIO",   "VALOR":MIN_LBS_NUEVA_MAQ_RATIO},
        {"KPI":"TAIL_MAX_LBS_KEY",          "VALOR":TAIL_MAX_LBS},
        {"KPI":"MAX_SEG_DIA",               "VALOR":MAX_SEG_DIA},
        {"KPI":"MAX_MAQ_POR_KEY",           "VALOR":MAX_MAQ_POR_KEY},
        {"KPI":"LIMITE_MAQ_DIA",            "VALOR":"SI" if max_machines else "NO"},
        {"KPI":"RANGO_PLAN",                "VALOR":f"{start_date.date()} → {end_date.date()}"},
    ])

    _dem_pend  = dem[dem["LBS_PENDIENTES"]>1e-9].copy()
    pend_by_tt = (_dem_pend.groupby(["DTITULAR","TIPO_TEJIDO"])["LBS_PENDIENTES"]
                  .sum().reset_index().rename(columns={"LBS_PENDIENTES":"LBS_PEND"}))
    used_machines = set(plan_full.loc[plan_full["LBS_ASIGNADAS"]>0,"MAQUINA"].unique())
    diag_rows = []
    for maq, info in compat_info.items():
        tit_fix   = norm_intlike(info.get("titular",""))
        tej_allow = sorted(list(info.get("allowed",set()))) or ["(ALL)"]
        usable_days = sum(1 for di in range(N) if (maq,di) not in blocked_day and hours_avail(maq,di)>0)
        pend = pend_by_tt.copy()
        if tit_fix:                pend = pend[pend["DTITULAR"]==tit_fix]
        if tej_allow != ["(ALL)"]: pend = pend[pend["TIPO_TEJIDO"].isin(set(tej_allow))]
        lbs_match = float(pend["LBS_PEND"].sum()) if not pend.empty else 0.0
        diag_rows.append({"MAQUINA":maq,"DTITULAR_FIJO":tit_fix,
                          "TIPOS_TEJIDO_PERMITIDOS":", ".join(tej_allow),
                          "DIAS_USABLES":usable_days,"LBS_PENDIENTES_MATCH":lbs_match,
                          "FUE_USADA":"SI" if maq in used_machines else "NO",
                          "ES_IDLE":"SI" if not machine_has_any_lbs(maq) else "NO"})
    diag_df = pd.DataFrame(diag_rows).sort_values(["FUE_USADA","LBS_PENDIENTES_MATCH","DIAS_USABLES"],
                                                    ascending=[True,False,False])

    pending_keys = dem[dem["LBS_PENDIENTES"]>1e-9].copy()
    if pending_keys.empty:
        dem_diag = pd.DataFrame(columns=["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO",
                                          "LBS_PENDIENTES","MAQS_COMPAT","MAQS_CON_DIA_LIBRE","MAQS_EJEMPLO"])
    else:
        gk = (pending_keys.groupby(["ESTILO_OPTIMO","LOTE_HILO","DTITULAR","TIPO_TEJIDO"])["LBS_PENDIENTES"]
              .sum().reset_index().sort_values("LBS_PENDIENTES", ascending=False))
        dd_rows = []
        for _, rr in gk.head(40).iterrows():
            elig = []; elig_free = []
            for maq in compat_info.keys():
                if machine_can(maq, rr["ESTILO_OPTIMO"], rr["DTITULAR"], rr["TIPO_TEJIDO"], rr["LOTE_HILO"]):
                    elig.append(maq)
                    if earliest_free_day(maq) is not None: elig_free.append(maq)
            dd_rows.append({"ESTILO_OPTIMO":rr["ESTILO_OPTIMO"],"LOTE_HILO":rr["LOTE_HILO"],
                            "DTITULAR":rr["DTITULAR"],"TIPO_TEJIDO":rr["TIPO_TEJIDO"],
                            "LBS_PENDIENTES":float(rr["LBS_PENDIENTES"]),
                            "MAQS_COMPAT":len(elig),"MAQS_CON_DIA_LIBRE":len(elig_free),
                            "MAQS_EJEMPLO":", ".join(elig_free[:20])})
        dem_diag = pd.DataFrame(dd_rows)

    return {
        "plan_full":      plan_full,
        "pivot_plan":     pivot_plan,
        "resumen":        resumen,
        "machines_day":   machines_day_df,
        "kpi":            kpi,
        "diag_maquinas":  diag_df,
        "diag_demanda":   dem_diag,
        "total_assigned": total_assigned,
        "lbs_pend_fin":   lbs_pend_fin,
        "cobertura":      round(total_assigned/(total_assigned+lbs_pend_fin)*100,1) if (total_assigned+lbs_pend_fin)>0 else 0,
    }


def build_excel(results):
    """Genera el Excel en memoria y retorna bytes."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        results["plan_full"].to_excel(w,     index=False, sheet_name="PLAN_DIARIO")
        results["pivot_plan"].to_excel(w,    index=False, sheet_name="PIVOT_PLAN")
        results["resumen"].to_excel(w,       index=False, sheet_name="RESUMEN")
        results["machines_day"].to_excel(w,  index=False, sheet_name="MACHINES_DAY")
        results["kpi"].to_excel(w,           index=False, sheet_name="KPI")
        results["diag_maquinas"].to_excel(w, index=False, sheet_name="DIAGNOSTICO_MAQUINAS")
        results["diag_demanda"].to_excel(w,  index=False, sheet_name="DIAGNOSTICO_DEMANDA")

    buf.seek(0)
    wb = load_workbook(buf)
    if "PIVOT_PLAN" in wb.sheetnames:
        ws = wb["PIVOT_PLAN"]
        col_maq = next((c for c in range(1, ws.max_column+1)
                        if str(ws.cell(1,c).value or "").strip().upper() == "MAQUINA"), 2)
        top_side = Side(style="thin")
        def add_top(cell):
            b = cell.border
            cell.border = Border(left=b.left, right=b.right, top=top_side, bottom=b.bottom)
        prev_maq = None
        for r in range(2, ws.max_row+1):
            mq = str(ws.cell(r, col_maq).value or "").strip()
            if prev_maq is not None and mq != prev_maq:
                for c in range(1, ws.max_column+1):
                    add_top(ws.cell(r, c))
            prev_maq = mq

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ============================================================
# STREAMLIT UI
# ============================================================
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── CSS personalizado ──
st.markdown("""
<style>
    /* Fondo general */
    .stApp { background-color: #0f1117; }

    /* Header principal */
    .main-header {
        background: linear-gradient(135deg, #1a1f2e 0%, #2d3561 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
        border: 1px solid #3d4f7c;
    }
    .main-header h1 { color: #e8eaf6; font-size: 2.2rem; margin: 0; }
    .main-header p  { color: #9fa8da; margin: 0.3rem 0 0; font-size: 0.95rem; }

    /* KPI cards */
    .kpi-card {
        background: linear-gradient(135deg, #1e2a3a 0%, #243447 100%);
        border: 1px solid #2d4a6b;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        text-align: center;
        transition: transform 0.2s;
    }
    .kpi-card:hover { transform: translateY(-2px); }
    .kpi-label { color: #8eb4d4; font-size: 0.78rem; font-weight: 600;
                 text-transform: uppercase; letter-spacing: 0.05em; }
    .kpi-value { color: #e8f4fd; font-size: 1.9rem; font-weight: 700; line-height: 1.2; }
    .kpi-sub   { color: #64b5f6; font-size: 0.82rem; margin-top: 0.2rem; }

    /* Cobertura especial */
    .kpi-coverage .kpi-value { color: #69f0ae; }

    /* Semáforo */
    .badge-green  { background:#1b5e20; color:#a5d6a7; padding:2px 10px;
                    border-radius:20px; font-size:0.8rem; font-weight:600; }
    .badge-yellow { background:#f57f17; color:#fff9c4; padding:2px 10px;
                    border-radius:20px; font-size:0.8rem; font-weight:600; }
    .badge-red    { background:#b71c1c; color:#ffcdd2; padding:2px 10px;
                    border-radius:20px; font-size:0.8rem; font-weight:600; }

    /* Sidebar */
    .sidebar-info { background:#1a2332; border:1px solid #2d4a6b;
                    border-radius:10px; padding:1rem; margin-top:0.5rem; }
    .sidebar-info p { color:#8eb4d4; margin:0.2rem 0; font-size:0.85rem; }

    /* Tab activo */
    .stTabs [data-baseweb="tab-list"] { background:#13171f; border-radius:10px; padding:4px; }
    .stTabs [data-baseweb="tab"] { border-radius:8px; color:#8eb4d4; }
    .stTabs [aria-selected="true"] { background:#2d3561 !important; color:#e8eaf6 !important; }

    /* Dataframes */
    .stDataFrame { border-radius: 10px; overflow: hidden; }

    /* Botón primario */
    .stButton>button[kind="primary"] {
        background: linear-gradient(135deg, #3d4f7c, #5c6bc0);
        border: none; border-radius: 10px; font-weight: 600;
        transition: all 0.2s;
    }
    .stButton>button[kind="primary"]:hover {
        background: linear-gradient(135deg, #5c6bc0, #7986cb);
        transform: translateY(-1px);
    }

    /* Download button */
    .stDownloadButton>button {
        background: linear-gradient(135deg, #1b5e20, #2e7d32);
        color: white; border: none; border-radius: 10px;
        font-weight: 600; font-size: 1rem; padding: 0.7rem;
        width: 100%; transition: all 0.2s;
    }
    .stDownloadButton>button:hover {
        background: linear-gradient(135deg, #2e7d32, #388e3c);
        transform: translateY(-1px);
    }

    /* Expander */
    .streamlit-expanderHeader { background:#1a2332; border-radius:8px; }

    /* Divider */
    hr { border-color: #2d4a6b; }
</style>
""", unsafe_allow_html=True)

# ── Paleta de colores para estilos/colores de hilo ──
COLOR_PALETTE = {
    "ASHEA":  "#FF8C00",
    "NATU":   "#4169E1",
    "WHITE":  "#F5F5F5",
    "BLACK":  "#444444",
    "RED":    "#DC143C",
    "BLUE":   "#1E90FF",
    "GREEN":  "#228B22",
    "YELLOW": "#FFD700",
    "GREY":   "#808080",
    "GRAY":   "#808080",
}

def get_color(val):
    if not val: return "#607d8b"
    v = str(val).upper().strip()
    for k, c in COLOR_PALETTE.items():
        if k in v: return c
    # Color determinístico por hash
    import hashlib
    h = int(hashlib.md5(v.encode()).hexdigest()[:6], 16)
    r = 80 + (h & 0xFF) % 120
    g = 80 + ((h >> 8) & 0xFF) % 120
    b = 120 + ((h >> 16) & 0xFF) % 120
    return f"#{r:02x}{g:02x}{b:02x}"

# ── Header ──
st.markdown("""
<div class="main-header">
    <h1>🧵 Plan Tejido v3</h1>
    <p>Motor de planificación con afinidad por color/yarn · DGREAL por máquina · Backfill inteligente</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ──
with st.sidebar:
    st.markdown("### 📁 Archivo de entrada")
    uploaded = st.file_uploader("Selecciona la plantilla Excel", type=["xlsx","xls"],
                                 help="Debe contener: PARAMETROS, REGLAS, ESTADO_MAQUINA, DEMANDA, COMPAT_MAQUINA, RATES, RESTRICCIONES")
    st.divider()

    if "results" in st.session_state:
        res = st.session_state["results"]
        st.markdown("### 📊 Estado del plan")
        cov = res["cobertura"]
        color_cov = "#69f0ae" if cov >= 90 else "#ffd740" if cov >= 70 else "#ff5252"
        st.markdown(f"""
        <div class="sidebar-info">
            <p>✅ <b>LBS Planificadas:</b> {res['total_assigned']:,.0f}</p>
            <p>⏳ <b>LBS Pendientes:</b> {res['lbs_pend_fin']:,.0f}</p>
            <p>🎯 <b>Cobertura:</b> <span style="color:{color_cov};font-weight:700">{cov}%</span></p>
        </div>
        """, unsafe_allow_html=True)
        st.divider()

    st.markdown("### 🔧 Filtros globales")
    filtro_tipo_tejido = st.multiselect("Tipo Tejido", options=[], key="global_tej",
                                         help="Se activa tras generar el plan")
    filtro_color = st.multiselect("Color", options=[], key="global_color",
                                   help="Se activa tras generar el plan")

    st.divider()
    st.markdown("""
    <div class="sidebar-info">
        <p>🔢 <b>Versión:</b> v3</p>
        <p>⚙️ <b>Motor:</b> 4 fases + backfill</p>
        <p>🎨 <b>Afinidad:</b> estilo → yarn → color</p>
        <p>📍 <b>DGREAL:</b> lookup por máquina</p>
    </div>
    """, unsafe_allow_html=True)

if uploaded is None:
    st.markdown("""
    <div style="background:#1a2332;border:1px solid #2d4a6b;border-radius:12px;
                padding:2.5rem;text-align:center;margin-top:1rem;">
        <div style="font-size:3rem">📂</div>
        <h3 style="color:#8eb4d4;margin:0.5rem 0">Sube tu plantilla Excel</h3>
        <p style="color:#607d8b;margin:0">Usa el panel lateral para cargar el archivo</p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Cargar hojas ──
@st.cache_data(show_spinner="Cargando hojas del Excel...")
def load_sheets(file_bytes):
    xls = io.BytesIO(file_bytes)
    def rt(sheet):   return pd.read_excel(xls, sheet_name=sheet, header=1)
    def ro(sheet, cols):
        try:    return pd.read_excel(xls, sheet_name=sheet, header=1)
        except: return pd.DataFrame(columns=cols)
    return {
        "params":   rt("PARAMETROS"),
        "reglas":   rt("REGLAS"),
        "estado":   rt("ESTADO_MAQUINA"),
        "demanda":  rt("DEMANDA"),
        "compat":   rt("COMPAT_MAQUINA"),
        "rates":    rt("RATES"),
        "restr":    rt("RESTRICCIONES"),
        "cal":      ro("CALENDARIO_MAQUINA", ["MAQUINA","FECHA_INICIO","FECHA_FIN","TIPO","HORAS_DISPONIBLES","NOTA"]),
    }

try:
    sheets = load_sheets(uploaded.read())
    st.success(f"✅ **{uploaded.name}** cargado correctamente")
except Exception as e:
    st.error(f"❌ Error cargando el archivo: {e}")
    st.stop()

# ── Vista previa de parámetros ──
with st.expander("👁️ Vista previa de parámetros", expanded=False):
    st.dataframe(sheets["params"], use_container_width=True, hide_index=True)

# ── Botón correr ──
st.divider()
col_btn, col_info = st.columns([1, 3])
with col_btn:
    run = st.button("▶️ Generar Plan", type="primary", use_container_width=True)
with col_info:
    if "results" in st.session_state:
        st.success("✅ Plan generado — puedes regenerarlo subiendo otro archivo o haciendo click en Generar Plan")

if not run and "results" not in st.session_state:
    st.stop()

if run:
    log_placeholder = st.empty()
    prog_bar = st.progress(0, text="Iniciando motor...")
    logs = []; counter = [0]; total_steps = 7
    def progress_cb(msg):
        logs.append(msg)
        counter[0] = min(counter[0] + 1, total_steps)
        pct = int(counter[0] / total_steps * 100)
        log_placeholder.markdown(f"```\n{msg}\n```")
        prog_bar.progress(pct, text=msg)

    with st.spinner(""):
        try:
            results = run_motor(
                params_tbl  = sheets["params"],
                reglas_tbl  = sheets["reglas"],
                estado_raw  = sheets["estado"],
                demanda_raw = sheets["demanda"],
                compat_raw  = sheets["compat"],
                rates_raw   = sheets["rates"],
                restr_raw   = sheets["restr"],
                cal_raw     = sheets["cal"],
                progress_cb = progress_cb,
            )
            st.session_state["results"] = results
            log_placeholder.empty()
            prog_bar.progress(100, text="✅ Completado")
        except Exception as e:
            st.error(f"❌ Error en el motor: {e}")
            st.exception(e)
            st.stop()

results = st.session_state["results"]
kpi_df  = results["kpi"]
def kv(name):
    row = kpi_df[kpi_df["KPI"]==name]
    return row.iloc[0]["VALOR"] if not row.empty else 0

# ── Actualizar filtros sidebar con opciones reales ──
plan_full = results["plan_full"]
plan_prod = plan_full[plan_full["LBS_ASIGNADAS"]>0]

# ── KPI Cards (HTML) ──
st.divider()
cov = results["cobertura"]
cov_color = "#69f0ae" if cov >= 90 else "#ffd740" if cov >= 70 else "#ff5252"
total_lbs  = results["total_assigned"]
pend_lbs   = results["lbs_pend_fin"]
h_setup    = kv("HORAS_SETUP_TOTALES")
h_prod     = kv("HORAS_PRODUCTIVAS_TOTALES")
n_maq_used = plan_prod["MAQUINA"].nunique()

c1, c2, c3, c4, c5, c6 = st.columns(6)
cards = [
    (c1, "LBS Planificadas",   f"{total_lbs:,.0f}",   f"de {total_lbs+pend_lbs:,.0f} totales", ""),
    (c2, "LBS Pendientes",     f"{pend_lbs:,.0f}",    "sin cubrir",                              ""),
    (c3, "Cobertura",          f"{cov}%",             "del total",                               "kpi-coverage"),
    (c4, "Horas Setup",        f"{h_setup:,.0f}",     "horas de cambio",                         ""),
    (c5, "Horas Productivas",  f"{h_prod:,.0f}",      "horas de producción",                     ""),
    (c6, "Máquinas Usadas",    f"{n_maq_used}",       "con producción",                          ""),
]
for col, label, val, sub, extra_cls in cards:
    with col:
        st.markdown(f"""
        <div class="kpi-card {extra_cls}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value" style="{'color:'+cov_color if extra_cls else ''}">{val}</div>
            <div class="kpi-sub">{sub}</div>
        </div>
        """, unsafe_allow_html=True)

# ── Tabs ──
st.divider()
tab0, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🏠 Dashboard", "📋 Plan Diario", "🗓️ Gantt & Heatmap",
    "📊 Pivot Plan", "📈 Resumen", "🔧 Máquinas", "📦 Demanda"
])

# ══════════════════════════════════════════════
# TAB 0 — DASHBOARD
# ══════════════════════════════════════════════
with tab0:
    st.markdown("### 📊 Dashboard General")
    col_g1, col_g2 = st.columns([1, 2])

    # Gauge de cobertura
    with col_g1:
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=results["cobertura"],
            delta={"reference": 100, "valueformat": ".1f"},
            number={"suffix": "%", "font": {"size": 48, "color": "#e8eaf6"}},
            gauge={
                "axis": {"range": [0, 100], "tickcolor": "#8eb4d4"},
                "bar":  {"color": "#5c6bc0"},
                "steps": [
                    {"range": [0,  70], "color": "#b71c1c"},
                    {"range": [70, 90], "color": "#f57f17"},
                    {"range": [90,100], "color": "#1b5e20"},
                ],
                "threshold": {"line": {"color": "#69f0ae", "width": 3},
                              "thickness": 0.8, "value": 100},
                "bgcolor": "#1a2332", "bordercolor": "#2d4a6b",
            },
            title={"text": "Cobertura del Plan", "font": {"color": "#8eb4d4", "size": 16}},
        ))
        fig_gauge.update_layout(paper_bgcolor="#13171f", font_color="#e8eaf6",
                                 height=300, margin=dict(t=60,b=20,l=30,r=30))
        st.plotly_chart(fig_gauge, use_container_width=True)

    # Donut LBS por color
    with col_g2:
        lbs_color = (plan_prod.groupby("COLOR")["LBS_ASIGNADAS"].sum()
                     .reset_index().sort_values("LBS_ASIGNADAS", ascending=False))
        lbs_color = lbs_color[lbs_color["COLOR"].str.strip() != ""]
        fig_donut = px.pie(
            lbs_color, values="LBS_ASIGNADAS", names="COLOR",
            title="Distribución LBS por Color de Hilo",
            color="COLOR",
            color_discrete_map={c: get_color(c) for c in lbs_color["COLOR"].unique()},
            hole=0.5,
        )
        fig_donut.update_traces(textposition="outside", textinfo="percent+label",
                                 textfont_color="#e8eaf6")
        fig_donut.update_layout(paper_bgcolor="#13171f", plot_bgcolor="#13171f",
                                 font_color="#e8eaf6", height=300,
                                 margin=dict(t=50,b=20,l=20,r=20),
                                 legend=dict(font_color="#8eb4d4"),
                                 showlegend=True)
        st.plotly_chart(fig_donut, use_container_width=True)

    # Barras LBS por día + máquinas activas
    st.markdown("#### 📅 Producción diaria")
    md = results["machines_day"]
    fig_day = make_subplots(specs=[[{"secondary_y": True}]])
    fig_day.add_trace(go.Bar(
        x=md["FECHA"].astype(str), y=md["LBS_TOTALES"],
        name="LBS Totales", marker_color="#5c6bc0", opacity=0.85,
    ), secondary_y=False)
    fig_day.add_trace(go.Scatter(
        x=md["FECHA"].astype(str), y=md["MAQUINAS_ACTIVAS"],
        name="Máquinas Activas", line=dict(color="#69f0ae", width=2),
        mode="lines+markers", marker=dict(size=6),
    ), secondary_y=True)
    fig_day.add_trace(go.Bar(
        x=md["FECHA"].astype(str), y=md["HORAS_SETUP_TOTALES"],
        name="Horas Setup", marker_color="#ff5252", opacity=0.6,
        visible="legendonly",
    ), secondary_y=False)
    fig_day.update_layout(
        paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
        font_color="#e8eaf6", height=350,
        legend=dict(orientation="h", y=1.08, font_color="#8eb4d4"),
        margin=dict(t=40,b=40,l=60,r=60),
        xaxis=dict(gridcolor="#2d4a6b", tickfont_color="#8eb4d4"),
        yaxis=dict(title="LBS", gridcolor="#2d4a6b", tickfont_color="#8eb4d4"),
        yaxis2=dict(title="Máquinas", tickfont_color="#69f0ae"),
        hovermode="x unified",
    )
    st.plotly_chart(fig_day, use_container_width=True)

    # Top máquinas por LBS
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        st.markdown("#### 🏭 Top máquinas por LBS")
        top_maq = (plan_prod.groupby("MAQUINA")["LBS_ASIGNADAS"].sum()
                   .reset_index().sort_values("LBS_ASIGNADAS", ascending=True).tail(15))
        fig_maq = px.bar(top_maq, x="LBS_ASIGNADAS", y="MAQUINA", orientation="h",
                          color="LBS_ASIGNADAS", color_continuous_scale="Blues",
                          labels={"LBS_ASIGNADAS":"LBS","MAQUINA":"Máquina"})
        fig_maq.update_layout(paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
                               font_color="#e8eaf6", height=420,
                               margin=dict(t=20,b=20,l=20,r=20),
                               coloraxis_showscale=False,
                               xaxis=dict(gridcolor="#2d4a6b"),
                               yaxis=dict(tickfont_color="#8eb4d4"))
        st.plotly_chart(fig_maq, use_container_width=True)

    with col_t2:
        st.markdown("#### 🎨 LBS por Estilo (Top 15)")
        top_est = (plan_prod.groupby("ESTILO_OPTIMO")["LBS_ASIGNADAS"].sum()
                   .reset_index().sort_values("LBS_ASIGNADAS", ascending=True).tail(15))
        fig_est = px.bar(top_est, x="LBS_ASIGNADAS", y="ESTILO_OPTIMO", orientation="h",
                          color="LBS_ASIGNADAS", color_continuous_scale="Purples",
                          labels={"LBS_ASIGNADAS":"LBS","ESTILO_OPTIMO":"Estilo"})
        fig_est.update_layout(paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
                               font_color="#e8eaf6", height=420,
                               margin=dict(t=20,b=20,l=20,r=20),
                               coloraxis_showscale=False,
                               xaxis=dict(gridcolor="#2d4a6b"),
                               yaxis=dict(tickfont_color="#8eb4d4"))
        st.plotly_chart(fig_est, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 1 — PLAN DIARIO
# ══════════════════════════════════════════════
with tab1:
    st.markdown("### 📋 Plan Diario")

    # Filtros en fila
    cf1, cf2, cf3, cf4 = st.columns(4)
    maqs_opts = ["(Todas)"] + sorted(plan_prod["MAQUINA"].astype(str).unique().tolist())
    ests_opts = ["(Todos)"] + sorted(plan_prod["ESTILO_OPTIMO"].astype(str).unique().tolist())
    col_opts  = ["(Todos)"] + sorted(plan_prod["COLOR"].dropna().astype(str).unique().tolist())
    tej_opts2 = ["(Todos)"] + sorted(plan_prod["TIPO_TEJIDO"].astype(str).unique().tolist())

    sel_maq = cf1.selectbox("🏭 Máquina", maqs_opts, key="pd_maq")
    sel_est = cf2.selectbox("🎽 Estilo",  ests_opts, key="pd_est")
    sel_col = cf3.selectbox("🎨 Color",   col_opts,  key="pd_col")
    sel_tej = cf4.selectbox("🧶 Tipo Tejido", tej_opts2, key="pd_tej")

    df_show = plan_prod.copy()
    if sel_maq != "(Todas)": df_show = df_show[df_show["MAQUINA"]==sel_maq]
    if sel_est != "(Todos)": df_show = df_show[df_show["ESTILO_OPTIMO"]==sel_est]
    if sel_col != "(Todos)": df_show = df_show[df_show["COLOR"]==sel_col]
    if sel_tej != "(Todos)": df_show = df_show[df_show["TIPO_TEJIDO"]==sel_tej]

    # Tabla con colores en COLOR y SETUP_TIPO
    def style_plan(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        if "COLOR" in df.columns:
            for i in df.index:
                c = get_color(df.at[i,"COLOR"])
                styles.at[i,"COLOR"] = f"background-color:{c}22;color:{c};font-weight:600"
        if "SETUP_TIPO" in df.columns:
            setup_colors = {
                "CAMBIO_TEJIDO": "#ff5252", "CAMBIO_ESTILO": "#ffd740",
                "CAMBIO_LOTE":   "#69b2ff", "CONTINUIDAD":   "#69f0ae",
                "INICIO":        "#b39ddb", "TAIL_SEED":     "#80cbc4",
                "TAIL_CONT":     "#80cbc4",
            }
            for i in df.index:
                st_val = str(df.at[i,"SETUP_TIPO"])
                c = setup_colors.get(st_val, "#8eb4d4")
                styles.at[i,"SETUP_TIPO"] = f"color:{c};font-weight:600"
        return styles

    cols_show = ["FECHA","MAQUINA","SEQ_MAQUINA","ESTILO_OPTIMO","ESTILO_REAL",
                 "LOTE_HILO","DTITULAR","DGREAL","TIPO_TEJIDO","YARN","COLOR",
                 "LBS_ASIGNADAS","HORAS_SETUP","HORAS_NETAS","RATE_LBS_DIA","SETUP_TIPO"]
    df_disp = df_show[[c for c in cols_show if c in df_show.columns]].copy()

    st.dataframe(
        df_disp.style.apply(style_plan, axis=None),
        use_container_width=True, hide_index=True, height=500,
    )
    st.caption(f"**{len(df_show):,}** segmentos · Total: **{df_show['LBS_ASIGNADAS'].sum():,.0f} LBS**")

    # Mini Gantt si hay filtro de máquina
    if sel_maq != "(Todas)" and len(df_show) > 0:
        st.markdown(f"#### 📅 Secuencia de {sel_maq}")
        gantt_data = df_show.sort_values("SEQ_MAQUINA").copy()
        gantt_data["Tarea"] = gantt_data["ESTILO_OPTIMO"] + " / " + gantt_data["LOTE_HILO"]
        fig_g = px.bar(
            gantt_data, x="LBS_ASIGNADAS", y="FECHA", orientation="h",
            color="COLOR", color_discrete_map={c: get_color(c) for c in gantt_data["COLOR"].unique()},
            hover_data=["ESTILO_OPTIMO","LOTE_HILO","YARN","SETUP_TIPO","LBS_ASIGNADAS"],
            labels={"LBS_ASIGNADAS":"LBS","FECHA":"Fecha"},
            title=f"Producción por día — Máquina {sel_maq}",
        )
        fig_g.update_layout(paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
                             font_color="#e8eaf6", height=max(250, len(df_show)*30),
                             margin=dict(t=50,b=30,l=80,r=30),
                             yaxis=dict(tickfont_color="#8eb4d4"),
                             xaxis=dict(gridcolor="#2d4a6b"),
                             legend_title_text="Color")
        st.plotly_chart(fig_g, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 2 — GANTT & HEATMAP
# ══════════════════════════════════════════════
with tab2:
    st.markdown("### 🗓️ Gantt — Máquinas × Días")

    col_g1, col_g2, col_g3 = st.columns(3)
    max_maqs = col_g1.slider("Máquinas a mostrar", 5, 60, 20, key="gantt_n")
    color_by = col_g2.selectbox("Colorear por", ["COLOR","ESTILO_OPTIMO","TIPO_TEJIDO","SETUP_TIPO"], key="gantt_cb")
    tej_f    = col_g3.selectbox("Tipo tejido", ["(Todos)"] + sorted(plan_prod["TIPO_TEJIDO"].unique().tolist()), key="gantt_tej")

    gdf = plan_prod.copy()
    if tej_f != "(Todos)": gdf = gdf[gdf["TIPO_TEJIDO"]==tej_f]

    # Tomar top N máquinas por LBS
    top_maqs = (gdf.groupby("MAQUINA")["LBS_ASIGNADAS"].sum()
                .nlargest(max_maqs).index.tolist())
    gdf = gdf[gdf["MAQUINA"].isin(top_maqs)].copy()
    gdf["FECHA_STR"] = gdf["FECHA"].astype(str)
    gdf["Info"] = (gdf["ESTILO_OPTIMO"] + " | " + gdf["LOTE_HILO"] +
                   " | " + gdf["LBS_ASIGNADAS"].apply(lambda x: f"{x:,.0f}") + " lbs")

    if gdf.empty:
        st.info("No hay datos para mostrar con los filtros seleccionados.")
    else:
        unique_vals = gdf[color_by].dropna().unique()
        color_map   = {v: get_color(v) for v in unique_vals}

        fig_gantt = px.bar(
            gdf.sort_values(["MAQUINA","FECHA"]),
            x="LBS_ASIGNADAS", y="MAQUINA", color=color_by,
            color_discrete_map=color_map, orientation="h",
            hover_data=["FECHA","ESTILO_OPTIMO","LOTE_HILO","YARN","SETUP_TIPO","LBS_ASIGNADAS"],
            labels={"LBS_ASIGNADAS":"LBS","MAQUINA":"Máquina"},
            title=f"Producción total por máquina — coloreado por {color_by}",
        )
        fig_gantt.update_layout(
            paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
            font_color="#e8eaf6",
            height=max(400, len(top_maqs) * 28),
            margin=dict(t=50,b=30,l=80,r=30),
            yaxis=dict(tickfont_color="#8eb4d4", categoryorder="total ascending"),
            xaxis=dict(gridcolor="#2d4a6b", title="LBS Totales"),
            legend=dict(font_color="#8eb4d4", title_font_color="#8eb4d4"),
            barmode="stack",
        )
        st.plotly_chart(fig_gantt, use_container_width=True)

    # Heatmap utilización
    st.markdown("#### 🔥 Heatmap — Utilización por máquina × día")
    heat_n = st.slider("Máquinas en heatmap", 5, 40, 15, key="heat_n")
    top_heat = (plan_prod.groupby("MAQUINA")["LBS_ASIGNADAS"].sum()
                .nlargest(heat_n).index.tolist())
    heat_df = (plan_prod[plan_prod["MAQUINA"].isin(top_heat)]
               .groupby(["MAQUINA","FECHA"])["LBS_ASIGNADAS"].sum().reset_index())
    heat_pivot = heat_df.pivot(index="MAQUINA", columns="FECHA", values="LBS_ASIGNADAS").fillna(0)

    if not heat_pivot.empty:
        fig_heat = px.imshow(
            heat_pivot,
            color_continuous_scale="Blues",
            labels=dict(x="Fecha", y="Máquina", color="LBS"),
            title="LBS por máquina y día",
            aspect="auto",
        )
        fig_heat.update_layout(
            paper_bgcolor="#13171f", plot_bgcolor="#13171f",
            font_color="#e8eaf6", height=max(300, heat_n * 25),
            margin=dict(t=50,b=40,l=80,r=20),
            xaxis=dict(tickfont_color="#8eb4d4", tickangle=-45),
            yaxis=dict(tickfont_color="#8eb4d4"),
        )
        st.plotly_chart(fig_heat, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 3 — PIVOT PLAN
# ══════════════════════════════════════════════
with tab3:
    st.markdown("### 📊 Pivot Plan")
    pivot = results["pivot_plan"]
    pf1, pf2 = st.columns(2)
    tej_piv = pf1.selectbox("Tipo Tejido", ["(Todos)"]+sorted(pivot["TIPO_TEJIDO"].astype(str).unique().tolist()), key="piv_tej")
    maq_piv = pf2.selectbox("Máquina",     ["(Todas)"]+sorted(pivot["MAQUINA"].astype(str).unique().tolist()), key="piv_maq")
    piv_show = pivot.copy()
    if tej_piv != "(Todos)": piv_show = piv_show[piv_show["TIPO_TEJIDO"]==tej_piv]
    if maq_piv != "(Todas)": piv_show = piv_show[piv_show["MAQUINA"].astype(str)==maq_piv]
    st.dataframe(piv_show, use_container_width=True, hide_index=True, height=500)
    st.caption(f"{len(piv_show):,} filas · Total general: {piv_show['Total_general'].sum():,.0f} LBS")

# ══════════════════════════════════════════════
# TAB 4 — RESUMEN
# ══════════════════════════════════════════════
with tab4:
    st.markdown("### 📈 Resumen por Key")
    res_df = results["resumen"].copy()
    r1, r2 = st.columns([2,1])
    exc_filter = r1.radio("Mostrar", ["Todos","Excedidos","Con diferencia","Sin cubrir"], horizontal=True)
    if exc_filter == "Excedidos":    res_df = res_df[res_df["EXCEDIDO"]=="SI"]
    elif exc_filter == "Con diferencia": res_df = res_df[res_df["DIFERENCIA"].abs() > 1]
    elif exc_filter == "Sin cubrir": res_df = res_df[res_df["LBS_ASIGNADAS"] < 1]

    # Colorear diferencia
    def style_resumen(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        if "DIFERENCIA" in df.columns:
            for i in df.index:
                v = df.at[i,"DIFERENCIA"]
                if v > 100:   styles.at[i,"DIFERENCIA"] = "color:#ff5252;font-weight:600"
                elif v < -100: styles.at[i,"DIFERENCIA"] = "color:#ffd740;font-weight:600"
                else:          styles.at[i,"DIFERENCIA"] = "color:#69f0ae"
        if "EXCEDIDO" in df.columns:
            for i in df.index:
                if df.at[i,"EXCEDIDO"] == "SI":
                    styles.at[i,"EXCEDIDO"] = "color:#ff5252;font-weight:700"
        return styles

    st.dataframe(
        res_df.style.apply(style_resumen, axis=None),
        use_container_width=True, hide_index=True, height=350,
    )

    # Gráfico plan vs asignado top 20
    top20 = results["resumen"].nlargest(20,"LBS_PLAN").copy()
    top20["KEY"] = top20["ESTILO_OPTIMO"].str[:10] + " / " + top20["LOTE_HILO"]
    fig_res = go.Figure()
    fig_res.add_trace(go.Bar(name="Plan", x=top20["KEY"], y=top20["LBS_PLAN"],
                              marker_color="#5c6bc0", opacity=0.8))
    fig_res.add_trace(go.Bar(name="Asignado", x=top20["KEY"], y=top20["LBS_ASIGNADAS"],
                              marker_color="#69f0ae", opacity=0.8))
    fig_res.update_layout(
        barmode="group", title="Top 20 keys: Plan vs Asignado",
        paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
        font_color="#e8eaf6", height=380,
        margin=dict(t=50,b=80,l=60,r=20),
        xaxis=dict(tickangle=-40, tickfont_color="#8eb4d4", gridcolor="#2d4a6b"),
        yaxis=dict(gridcolor="#2d4a6b"),
        legend=dict(font_color="#8eb4d4"),
    )
    st.plotly_chart(fig_res, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 5 — DIAGNÓSTICO MÁQUINAS
# ══════════════════════════════════════════════
with tab5:
    st.markdown("### 🔧 Diagnóstico de Máquinas")
    dm = results["diag_maquinas"].copy()

    col_d1, col_d2 = st.columns(2)
    idle_only   = col_d1.checkbox("Solo máquinas ociosas")
    no_usadas   = col_d2.checkbox("Solo máquinas no usadas")
    if idle_only: dm = dm[dm["ES_IDLE"]=="SI"]
    if no_usadas: dm = dm[dm["FUE_USADA"]=="NO"]

    # Semáforo visual
    def style_diag(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for i in df.index:
            usada = df.at[i,"FUE_USADA"] if "FUE_USADA" in df.columns else "NO"
            idle  = df.at[i,"ES_IDLE"]   if "ES_IDLE"   in df.columns else "SI"
            lbs   = float(df.at[i,"LBS_PENDIENTES_MATCH"]) if "LBS_PENDIENTES_MATCH" in df.columns else 0
            if usada == "SI" and idle == "NO":
                c = "#69f0ae"  # verde
            elif usada == "SI" and idle == "SI":
                c = "#ffd740"  # amarillo (usada pero ociosa — raro)
            elif lbs > 0:
                c = "#ffd740"  # amarillo — no usada pero tiene demanda compatible
            else:
                c = "#ff5252"  # rojo — no usada, sin demanda
            for col in ["FUE_USADA","ES_IDLE"]:
                if col in df.columns:
                    styles.at[i,col] = f"color:{c};font-weight:700"
        return styles

    st.dataframe(
        dm.style.apply(style_diag, axis=None),
        use_container_width=True, hide_index=True, height=450,
    )

    # Resumen semáforo
    total_maq  = len(results["diag_maquinas"])
    usadas     = (results["diag_maquinas"]["FUE_USADA"]=="SI").sum()
    ociosas    = (results["diag_maquinas"]["ES_IDLE"]=="SI").sum()
    sc1, sc2, sc3 = st.columns(3)
    sc1.markdown(f'<div style="text-align:center"><span class="badge-green">✅ Usadas: {usadas}</span></div>', unsafe_allow_html=True)
    sc2.markdown(f'<div style="text-align:center"><span class="badge-yellow">⚠️ Ociosas: {ociosas}</span></div>', unsafe_allow_html=True)
    sc3.markdown(f'<div style="text-align:center"><span class="badge-red">❌ Sin usar: {total_maq-usadas}</span></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════
# TAB 6 — DIAGNÓSTICO DEMANDA
# ══════════════════════════════════════════════
with tab6:
    st.markdown("### 📦 Diagnóstico de Demanda Pendiente")
    dd = results["diag_demanda"]
    if dd.empty:
        st.markdown("""
        <div style="background:#1b5e20;border:1px solid #2e7d32;border-radius:12px;
                    padding:2rem;text-align:center;">
            <div style="font-size:2.5rem">🎉</div>
            <h3 style="color:#a5d6a7">¡Toda la demanda fue cubierta!</h3>
            <p style="color:#81c784">0 LBS pendientes sin asignar</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.warning(f"⚠️ **{len(dd)}** keys con demanda pendiente (top 40)")
        def style_demanda(df):
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            if "MAQS_COMPAT" in df.columns:
                for i in df.index:
                    mc = int(df.at[i,"MAQS_COMPAT"]) if df.at[i,"MAQS_COMPAT"] else 0
                    mf = int(df.at[i,"MAQS_CON_DIA_LIBRE"]) if df.at[i,"MAQS_CON_DIA_LIBRE"] else 0
                    c = "#69f0ae" if mf > 0 else "#ffd740" if mc > 0 else "#ff5252"
                    styles.at[i,"MAQS_COMPAT"]     = f"color:{c};font-weight:600"
                    styles.at[i,"MAQS_CON_DIA_LIBRE"] = f"color:{c};font-weight:600"
            return styles
        st.dataframe(
            dd.style.apply(style_demanda, axis=None),
            use_container_width=True, hide_index=True, height=400,
        )
        # Gráfico top pendientes
        top_pend = dd.nlargest(15,"LBS_PENDIENTES").copy()
        top_pend["KEY"] = top_pend["ESTILO_OPTIMO"].str[:10] + "/" + top_pend["LOTE_HILO"]
        fig_pend = px.bar(top_pend, x="LBS_PENDIENTES", y="KEY", orientation="h",
                           color="MAQS_CON_DIA_LIBRE",
                           color_continuous_scale=["#b71c1c","#f57f17","#1b5e20"],
                           labels={"LBS_PENDIENTES":"LBS Pendientes","KEY":"Key","MAQS_CON_DIA_LIBRE":"Máqs libres"},
                           title="Top 15 keys pendientes (color = máquinas disponibles)")
        fig_pend.update_layout(paper_bgcolor="#13171f", plot_bgcolor="#1a2332",
                                font_color="#e8eaf6", height=420,
                                margin=dict(t=50,b=20,l=20,r=20),
                                xaxis=dict(gridcolor="#2d4a6b"),
                                yaxis=dict(tickfont_color="#8eb4d4",categoryorder="total ascending"))
        st.plotly_chart(fig_pend, use_container_width=True)

# ══════════════════════════════════════════════
# TAB 8 — DESCARGAR
# ══════════════════════════════════════════════
# (usa el mismo índice del tabs aunque visualmente es tab7)
st.divider()
st.markdown("### ⬇️ Descargar Excel completo")
col_dl1, col_dl2 = st.columns([2,1])
with col_dl1:
    st.markdown("""
    El archivo incluye **7 hojas**:
    - **PLAN_DIARIO** — todos los segmentos con SEQ_MAQUINA, COLOR, YARN, DGREAL
    - **PIVOT_PLAN** — vista cruzada máquina × fecha con líneas separadoras
    - **RESUMEN** — LBS plan vs asignado por key
    - **MACHINES_DAY** — actividad diaria
    - **KPI** — indicadores del plan
    - **DIAGNOSTICO_MAQUINAS** — utilización por máquina
    - **DIAGNOSTICO_DEMANDA** — keys pendientes
    """)
with col_dl2:
    with st.spinner("Preparando Excel..."):
        excel_bytes = build_excel(results)
    st.download_button(
        label="📥 Descargar PLAN_TEJIDO_v3.xlsx",
        data=excel_bytes,
        file_name="PLAN_TEJIDO_v3.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
